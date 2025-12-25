from datetime import datetime
import time
import glob
import os

from openpyxl import load_workbook
import pandas as pd
import pyodbc


def sum_cells(ws, cells):
    """Safely sum a list of cell references, treating None as 0."""
    return sum(ws[cell].value or 0 for cell in cells)


def run_bed_report_etl(
    sql_connection_string,
    template_path,
    external_folder_1,
    output_folder,
):
    """
    Example ETL:
    - Load latest Excel files from one folder
    - Run a SQL query to get metric totals
    - Combine everything into a single submission workbook
    """

    start_time = time.time()

    # Connect to SQL
    conn = pyodbc.connect(sql_connection_string)

    # Find latest Excel files from two directories
    files_1 = glob.glob(os.path.join(external_folder_1, "*.xlsx"))


    if not files_1:
        print("External files not found; check folder paths.")
        return

    latest_file_1 = max(files_1, key=os.path.getctime)


    # Load workbooks
    wb = load_workbook(template_path)
    wb_ext_1 = load_workbook(latest_file_1)

    ws = wb.active
    ws_ext_1 = wb_ext_1.active


    # Example generic SQL extracting several metrics
    query = """
    DECLARE @Yesterday date = DATEADD(day, -1, CAST(GETDATE() as date));
    DECLARE @Window7 date = DATEADD(day, -6, @Yesterday);
    DECLARE @Window14 date = DATEADD(day, -13, @Yesterday);
    DECLARE @Window21 date = DATEADD(day, -20, @Yesterday);

    -- Admissions
    SELECT 'Admissions' AS Metric, COUNT(*) AS Total
    FROM MainDB.FactTable FT
    WHERE FT.AdmissionDate = @Yesterday

    UNION ALL

    -- Emergency Admissions
    SELECT 'EmergencyAdmissions' AS Metric, COUNT(*) AS Total
    FROM MainDB.FactTable FT
    WHERE FT.AdmissionDate = @Yesterday
      AND FT.AdmissionType = 'Emergency'

    UNION ALL

    -- Discharges
    SELECT 'Discharges' AS Metric, COUNT(*) AS Total
    FROM MainDB.FactTable FT
    WHERE FT.DischargeDate = @Yesterday

    UNION ALL

    -- 7+ Day Stays
    SELECT 'Stay7Plus' AS Metric, COUNT(*) AS Total
    FROM MainDB.FactTable FT
    WHERE FT.AdmissionDate <= @Window7
      AND FT.DischargeDate IS NULL

    UNION ALL

    -- 14+ Day Stays
    SELECT 'Stay14Plus' AS Metric, COUNT(*) AS Total
    FROM MainDB.FactTable FT
    WHERE FT.AdmissionDate <= @Window14
      AND FT.DischargeDate IS NULL

    UNION ALL

    -- 21+ Day Stays
    SELECT 'Stay21Plus' AS Metric, COUNT(*) AS Total
    FROM MainDB.FactTable FT
    WHERE FT.AdmissionDate <= @Window21
      AND FT.DischargeDate IS NULL;
    """

    df_metrics = pd.read_sql(query, conn)

    # Example cell ranges in external workbook #1:
    bed_base_cells = ["C22", "C25", "C28", "C31"]
    beds_closed_cells = ["E22", "E25", "E28", "E31"]
    beds_occupied_cells = ["I22", "I25", "I28", "I31"]

    # Example summary metrics from workbook #2:
    # (Adjust these references to your anonymised template)
    covid_cells = ["D18", "I18", "M18"]

    # Fill in template with dates + metrics (replace cell refs for your layout)
    ws["B7"].value = datetime.today().strftime("%d-%b")

    ws["B25"].value = df_metrics.loc[
        df_metrics["Metric"] == "Admissions", "Total"
    ].values[0]
    ws["B26"].value = df_metrics.loc[
        df_metrics["Metric"] == "EmergencyAdmissions", "Total"
    ].values[0]
    ws["B28"].value = df_metrics.loc[
        df_metrics["Metric"] == "Discharges", "Total"
    ].values[0]

    # Example calculations from external workbook 1
    core_beds_open = sum_cells(ws_ext_1, bed_base_cells) - sum_cells(
        ws_ext_1, beds_closed_cells
    )
    total_beds_occupied = sum_cells(ws_ext_1, beds_occupied_cells)

    ws["B44"].value = core_beds_open
    ws["B46"].value = core_beds_open
    ws["B47"].value = total_beds_occupied

    # Example aggregated counts from external workbook 2
    ws["B81"].value = sum_cells(ws_ext_1, covid_cells)
    ws["B82"].value = ws["B81"].value

    # Long-stay metrics from the SQL result
    ws["B86"].value = df_metrics.loc[
        df_metrics["Metric"] == "Stay7Plus", "Total"
    ].values[0]
    ws["B87"].value = df_metrics.loc[
        df_metrics["Metric"] == "Stay14Plus", "Total"
    ].values[0]
    ws["B88"].value = df_metrics.loc[
        df_metrics["Metric"] == "Stay21Plus", "Total"
    ].values[0]

    # Save with dated filename
    os.makedirs(output_folder, exist_ok=True)
    output_path = os.path.join(
        output_folder, f"Submission_{datetime.now().strftime('%Y-%m-%d')}.xlsx"
    )
    wb.save(output_path)

    elapsed = round(time.time() - start_time, 2)
    print(f"ETL completed and saved to {output_path} ({elapsed} seconds).")

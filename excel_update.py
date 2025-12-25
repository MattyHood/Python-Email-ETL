import os
import time
from datetime import datetime
import glob

from openpyxl import load_workbook
from openpyxl.styles import NamedStyle
import comtypes.client


def recalculate_excel(file_path):
    """
    Opens an Excel file via COM, refreshes formulas, and saves.
    """
    excel = comtypes.client.CreateObject("Excel.Application")
    excel.Visible = False
    wb = excel.Workbooks.Open(file_path)
    wb.RefreshAll()
    excel.CalculateUntilAsyncQueriesDone()
    wb.Save()
    wb.Close()
    excel.Quit()


def update_template_with_latest_data(
    template_path, source_folder, pattern="*.xlsx", insert_rows=10):
    """
    Example: Load a template workbook, insert rows, copy values from the latest
    source workbook, and apply formulas/formatting.
    """

    start_time = time.time()
    today = datetime.today().date()

    # Find latest source file
    file_list = glob.glob(os.path.join(source_folder, pattern))
    if not file_list:
        print("No source files found.")
        return

    latest_source = max(file_list, key=os.path.getmtime)

    wb = load_workbook(template_path)
    wb_src = load_workbook(latest_source)

    ws = wb.active
    ws_src = wb_src.active

    # Insert blank rows at top to add new data
    ws.insert_rows(idx=2, amount=insert_rows)

    # Define some named styles
    styles = {
        "percent_style": "0.00%",
        "number_style": "0.00",
        "date_style": "DD/MM/YYYY",
    }

    def create_named_style(name, number_format):
        if name not in wb.named_styles:
            style = NamedStyle(name=name, number_format=number_format)
            wb.add_named_style(style)

    for style_name, fmt in styles.items():
        create_named_style(style_name, fmt)

    # Example: apply styles to some columns (adjust as needed)
    for row in range(2, ws.max_row + 1):
        ws[f"P{row}"].style = "percent_style"
        ws[f"R{row}"].style = "percent_style"
        ws[f"N{row}"].style = "number_style"
        ws[f"O{row}"].style = "number_style"
        ws[f"Q{row}"].style = "number_style"

    # Example: fill formulas and copy values from the external workbook
    # (Replace columns here to match your real template structure)
    for row in range(2, 12):
        ws[f"A{row}"].value = f'=IF(B{row}=TODAY(),"Y","N")'
        ws[f"B{row}"].value = today
        ws[f"B{row}"].style = "date_style"

        # Example of pulling data from another file
        ws[f"D{row}"].value = ws_src[f"B{row}"].value
        ws[f"N{row}"].value = ws_src[f"C{row}"].value
        ws[f"O{row}"].value = ws_src[f"D{row}"].value
        ws[f"Q{row}"].value = ws_src[f"E{row}"].value

    # Example of summary row
    summary_row = 12
    ws[f"D{summary_row}"].value = "Total"
    ws[f"N{summary_row}"].value = sum(
        ws[f"N{row}"].value or 0 for row in range(2, summary_row)
    )
    ws[f"O{summary_row}"].value = sum(
        ws[f"O{row}"].value or 0 for row in range(2, summary_row)
    )
    ws[f"Q{summary_row}"].value = sum(
        ws[f"Q{row}"].value or 0 for row in range(2, summary_row)
    )

    # Save and recalc formulas via Excel
    wb.save(template_path)
    recalculate_excel(template_path)

    elapsed = round(time.time() - start_time, 2)
    print(f"Excel template updated and recalculated in {elapsed} seconds.")
